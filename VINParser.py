import asyncio
from datetime import datetime
import itertools
import os
import random
import aiohttp
import openpyxl
from openpyxl import Workbook
import pandas as pd
from tqdm import tqdm


class AutodocParser:
    def __init__(self):
        self.today = datetime.now().strftime("%d/%m-%H.%M")
        self.now = datetime.now().strftime("%H.%M")
        self.inputDirectoryPath = "input/"
        self.outputDirectoryPath = "results/"
        self.inputFilePath = self.inputDirectoryPath + "vins.xlsx"
        self.outputFilePath = self.outputDirectoryPath + f"vin_{self.now}.xlsx"
        self.accountsFilePath = "accounts.xlsx"
        self.accountsHeaders = ["login", "password"]
        self.data = []
        
        self.timeout = 0
        self.sessionTimeout = aiohttp.ClientTimeout(total=None, sock_connect=self.timeout, sock_read=self.timeout)
        self.connector = aiohttp.TCPConnector(ssl=False, limit=100)
        self.session = aiohttp.ClientSession(connector=self.connector, timeout=self.sessionTimeout)

        
    def run(self):
        self.accounts = self.readAccountsFile()
        self.account = random.choice(self.accounts)
        self.tokenPostData = {
            "username": self.account["login"],
            "password": self.account["password"],
            "grant_type": "password"
        }
        self.data = self.readInputFile()
        if len(self.data) == 1: return
        loop = asyncio.get_event_loop()
        loop.run_until_complete(self.startParsing())


    async def startParsing(self):
        requestsList = []
        # * Запросы по VIN
        for vin in self.data:
            request = self.makePrimaryRequest(vin)
            requestsList.append(request)
        responses = await asyncio.gather(*requestsList)
        requestsList.clear()

        # *  Основная информация по машинам
        primaryData = []
        for i, response in enumerate(responses):
            primaryData.append(self.collectPrimaryData(response, self.data[i]))
        carsPrimaryInfo = await asyncio.gather(*primaryData)

        # * Парсинг по VIN по машине
        requestsList = []
        for carData in carsPrimaryInfo:
            requestsList.append(self.parseVINs(carData))
        carsInfo = await asyncio.gather(*requestsList)
        self.session.close()
        self.writeToExcel(carsInfo)

        await self.session.close()


    async def parseVINs(self, carInfo):
        if not carInfo: return
        detailsData = []
        uniqueParts = []

        carID = carInfo.get("CarID", "0")
        carModel = carInfo.get("Model", "")
        vin = carInfo.get("VIN", "")

        categoriesUrl = f"https://catalogoriginal.autodoc.ru/api/catalogs/original/brands/{carInfo['Catalog']}/cars/{carID}/categories?ssd={carInfo['Ssd']}".strip()
        response = await self.session.get(categoriesUrl)
        carSparePartsCategories = await response.json()

        categories = carSparePartsCategories.get("items", []) if "items" in carSparePartsCategories.keys() else carSparePartsCategories
        
        # Получение всех подкатегорий
        subcategories = [self.getChildrenList(category.get("children", [])) for category in categories]
        subcats = list(itertools.chain(*subcategories))
        allCategories = categories + subcats

        for partCategory in tqdm(allCategories):
            sparePartInfoUrl = f"https://catalogoriginal.autodoc.ru/api/catalogs/original/brands/{carInfo['Catalog']}/cars/0/categories/{partCategory['categoryId']}/units?ssd={partCategory['ssd']}"
            sparePartInfoResponse = await self.session.get(sparePartInfoUrl)
            sparePartInfoResponseJson = await sparePartInfoResponse.json()

            for sparePartsInfo in sparePartInfoResponseJson.get("items", []):
                sparePartDetailInfoUrl = f"https://catalogoriginal.autodoc.ru/api/catalogs/original/brands/{carInfo['Catalog']}/cars/0/units/{sparePartsInfo['unitId']}/spareparts?ssd={sparePartsInfo['ssd']}"
                data = {
                    "Ssd": sparePartsInfo['ssd']
                }
                sparePartDetailInfoResponse = await self.session.post(sparePartDetailInfoUrl, data=data)
                sparePartDetailInfoResponseJson = await sparePartDetailInfoResponse.json()
                spareParts = sparePartDetailInfoResponseJson.get("items", [])
                

                for sparePart in spareParts:
                    # Фильтр дубликатов #! Возможно из-за этого будут проблемы, могут попросить оставить дубли деталей
                    if sparePart["partNumber"] not in uniqueParts:
                        uniqueParts.append(sparePart["partNumber"])
                        
                        detailsData.append({
                            "category": partCategory["name"],
                            "partName": sparePart["name"],
                            "partNumber": sparePart["partNumber"]
                        })
                    else:
                        continue

        outputData = {
            "name": carInfo['Catalog'],
            "carModel": carModel,
            "VIN": vin,
            "items": detailsData
        }
        return outputData


    def getChildrenList(self, node):
        items = []
        for item in node:
            if isinstance(item, str): return items
            children = item.get("children", {})
            if children:
                items = self.getChildrenList(item)
            else:
                items.append(item)
        return items


    # Обычные запросы по VIN
    async def collectPrimaryData(self, response, vin):
        carPrimaryInfo = {
            "VIN": vin
        }
        responseData = await response.json()

        try:
            primaryData = responseData["commonAttributes"]
        except Exception as error:
            # print(error)
            return

        for attribute in primaryData:
            carPrimaryInfo[attribute["key"]] = attribute["value"]
        return carPrimaryInfo


    async def makePrimaryRequest(self, vin):
        clientId = random.randint(0, 500)
        # clientId = 333

        keyUrl = f"https://catalogoriginal.autodoc.ru/api/catalogs/original/cars/{vin}/modifications?clientId={clientId}"
        response = await self.session.get(keyUrl)
        return response

    
    def readAccountsFile(self):
        rawData = pd.read_excel("accounts.xlsx", index_col=False)
        accounts = rawData.to_numpy()
        
        accountsData = []
        for account in accounts:
            accountsData.append({
                "login": account[0],
                "password": account[1]
            })
        return accountsData


    def writeToExcel(self, carsData):
        headers = ["Категория", "Название запчасти", "Каталожный номер"]
        if not self.data: return
        writer = pd.ExcelWriter(self.outputFilePath, engine='openpyxl')
        #! car = [{...}, {...}, ...]
        for car in carsData:
            if not car: continue
            items = car["items"]
            sheetName = f"{car['name']} {car['VIN']}"
            df = pd.DataFrame(items)
            df.to_excel(writer, sheet_name=sheetName, encoding="utf-8", header=headers, index=False)
            writer.save()
        writer.close()


    def readInputFile(self):
        vins = pd.read_excel(self.inputFilePath, header=None)
        data = [vin[0] for vin in vins.values.tolist()]
        return data
    

    def setup(self):
        isFirstLaunch = False

        # input dir
        if not os.path.exists(self.inputDirectoryPath):
            os.mkdir(self.inputDirectoryPath)
            isFirstLaunch = True

        # input file
        if not os.path.exists(self.inputFilePath):
            wb = Workbook()
            wb.worksheets[0].cell(1 , 1, value="VINS")
            wb.save(self.inputFilePath)
            isFirstLaunch = True

        # output dir
        if not os.path.exists(self.outputDirectoryPath):
            os.mkdir(self.outputDirectoryPath)
            isFirstLaunch = True

        # accounts file
        if not os.path.exists(self.accountsFilePath):
            wb = Workbook()
            for i, header in enumerate(self.accountsHeaders):
                wb.worksheets[0].cell(1 , i+1, value=header)
            wb.save(self.accountsFilePath)
            isFirstLaunch = True
        return isFirstLaunch


if __name__ == "__main__":
    
    autodocParser = AutodocParser()
    if not autodocParser.setup():
        autodocParser.run()
    else:
        print("Рабочие файлы и папки созданы, заполните их!")
