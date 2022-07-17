import asyncio
from datetime import datetime
import os
from pprint import pprint
import random
import openpyxl
from openpyxl import Workbook
import pandas as pd

import aiohttp
from fake_useragent import UserAgent
from tqdm import tqdm

class DetailParser:
    def __init__(self):
        self.today = datetime.now().strftime("%d/%m-%H.%M")
        self.now = datetime.now().strftime("%H.%M")
        self.inputDirectoryPath = "input/"
        self.outputDirectoryPath = "results/"
        self.inputFilePath = self.inputDirectoryPath + "details.xlsx"
        self.outputFilePath = self.outputDirectoryPath + f"details_{self.now}.xlsx"
        self.accountsFilePath = "accounts.xlsx"
        self.accountsHeaders = ["login", "password"]
        self.inputHeaders = ["Наименование номенклатуры", "Каталожный номер"]
        
        self.analogOutputHeaders = ["Тип", "Наименование номенклатуры", "Каталожный номер", "Минимальный срок доставки", "Производитель", "Цена"]
        self.originalOutputHeaders = ["Тип", "Наименование номенклатуры", "Каталожный номер", "Срок доставки", "Минимальный срок доставки", "Производитель", "Количество", "Цена", "Поставщик"]

        
        self.timeout = 0
        self.sessionTimeout = aiohttp.ClientTimeout(total=None, sock_connect=self.timeout, sock_read=self.timeout)
        self.connector = aiohttp.TCPConnector(ssl=False, limit=100)
        self.userAgent = UserAgent().random

        self.loginAttempts = ["DC1/O1127x9ZL4GU2bhQgg==", "W7F+x+sPZUPsCAcXwYSH5Q=="]        
        


    async def startParsing(self):
        challengeGuid = await self.getChallengeGuid()
        tokenData = await self.getToken()
        self.outputHeaders = {
            "authorization": tokenData["token_type"] + " " + tokenData["access_token"],
            "user-agent": self.userAgent
        }
        loginData = await self.getLoginData(challengeGuid, random.choice(self.loginAttempts))
        session = loginData["session"]
        if loginData["response"]["clientStatus"] != 0:
            await self.startParsing()
            return
        # profileData = await self.getProfileData()

        # Проход по номерам деталей
        outputItems = []
        requests = []
        for detailData in self.data:
            requests.append(self.parseDetail(detailData, session, outputItems))
        responses = [await detail for detail in tqdm(asyncio.as_completed(requests), total=len(requests))]
        await session.close()

        self.writeToExcel(outputItems)


    async def parseDetail(self, detailData, session, outputItems):
        detailName = detailData["detailName"]
        detailNumber = detailData["detailNumber"]
        manufacturers = await self.getManufacturerInfo(detailNumber.replace("-", ""), session)
        
        # Проход по номерам производителей для каждой детали
        for manufacturer in manufacturers:
            manufacturerID = manufacturer["id"]
            manufacturerName = manufacturer["manufacturerName"]
            partName = manufacturer["partName"]
            partNumber = manufacturer["artNumber"]

            #* Вся инфа о детали
            detailInfo = await self.getDetailInfo(session, manufacturerID, partNumber)
            originalDetails = detailInfo["originals"]
            analogDetails = detailInfo["analogs"]

            #* Оригиналы
            for item in originalDetails["inventoryItems"]:
                outputDetailData = {
                    "type": "original",
                    "detailName": detailName,
                    "detailNumber": detailNumber,
                    "deliveryDays": item["deliveryDays"],
                    "minimumDeliveryDays": item["minimalDeliveryDays"],
                    "manufacturerName": manufacturerName,
                    "quantity": item["quantity"],
                    "price": item["price"],
                    "supplier": item["supplier"]["name"] # Поставщик
                }
                outputItems.append(outputDetailData)

            #* Аналоги
            for item in analogDetails["analogs"]:
                outputDetailData = {
                    "type": "analog",
                    "detailName": detailName,
                    "detailNumber": item["partNumber"], # можно использовать displayPartNumber
                    "minimumDeliveryDays": item["minimalDeliveryDays"],
                    "manufacturerName": item["manufacturer"]["name"],
                    "price": item["minimalPrice"],
                }
                outputItems.append(outputDetailData)


    async def getDetailInfo(self, session, manufacturerID, detailNumber):
        originalsUrl = f"https://webapi.autodoc.ru/api/spareparts/{manufacturerID}/{detailNumber}/2?framesId=undefined&attempt=undefined&isrecross=false"
        analogsUrl = f"https://webapi.autodoc.ru/api/spareparts/analogs/{manufacturerID}/{detailNumber}/2"
        session.headers["hash_"] = await self.getHash(manufacturerID, detailNumber)       
        session.headers["dnt"] = "1"       
        session.headers["source_"] = "Site2"       
        
        # original
        originalsResponse = await session.get(originalsUrl)
        originalsResponseJson = await originalsResponse.json()

        # analogs
        analogsResponse = await session.get(analogsUrl)
        analogsResponseJson = await analogsResponse.json()

        returnData = {
            "originals": originalsResponseJson,
            "analogs": analogsResponseJson
        }

        return returnData


    async def getManufacturerInfo(self, partNumber, session):
        url = f"https://webapi.autodoc.ru/api/manufacturers/{partNumber}?showAll=false"
        response = await session.get(url)
        responseJson = await response.json()
        return responseJson

    
    async def getProfileData(self):
        url = "https://webapi.autodoc.ru/api/client/profile"
        async with aiohttp.request("GET", url) as response:
            responseJson = await response.json()
        return responseJson


    async def getLoginData(self, challengeGuid, attempt):
        url = "https://webapi.autodoc.ru/api/account/login"
        
        data = {
            "attempt": attempt,
            "challengeGuid": challengeGuid,
            "gRecaptchaResponse": "",
            "login": self.account["login"],
            "password": self.account["password"],
            "rememberMe": "true"
        }

        session = aiohttp.ClientSession(connector=self.connector, headers=self.outputHeaders)
        response = await session.post(url, data=data)
        responseJson = await response.json()
        returnData = {
            "response": responseJson,
            "session": session
        }
        
        return returnData


    async def createSession(self, authData, hash) -> aiohttp.ClientSession:
        headers = {
            "user-agent": self.userAgent,
            "authorization": authData["token_type"] + " " + authData["access_token"],
            "hash": hash
        }
        session = aiohttp.ClientSession(connector=self.connector, timeout=self.sessionTimeout, headers=headers)
        return session


    #! Тут может вылезти капча
    async def getChallengeGuid(self):
        url = "https://webapi.autodoc.ru/api/captha?resource=Auth"
        async with aiohttp.request("GET", url) as response:
            responseJson = await response.json()
            challengeGuid = responseJson["challengeGuid"]
        return challengeGuid


    async def getHash(self, manufacturerID, partNumber):
        hashUrl = f"https://webapi.autodoc.ru/api/spareparts/hash/{manufacturerID}/{partNumber}"
        async with aiohttp.request("POST", hashUrl) as response:
            responseJson = await response.json()
        return responseJson


    #! За большое число запросов банят аккаунт нахуй
    async def getToken(self):
        responseJson = await self.makeTokenRequest()

        #! Если забанят аккаунт, берём другой рандомный
        while responseJson.get("error", False):
            if responseJson["error"] == "access_denied":
                print(f"Аккаунт {self.account['login']} недействителен")
            print(f"Аккаунт {self.account['login']} забанен, пробую ещё...")
            
            responseJson = await self.makeTokenRequest()
        
        print(f"Зашёл через аккаунт {self.account['login']}")
        accessToken = responseJson["access_token"]
        refreshToken = responseJson["refresh_token"]
        expiresIn = responseJson["expires_in"]
        tokenType = responseJson["token_type"]
    
        authData = {
            "access_token": accessToken,
            "refresh_token": refreshToken,
            "expires_in": expiresIn,
            "token_type": tokenType,
        }
        return authData 


    async def makeTokenRequest(self):
        url = "https://auth.autodoc.ru/token"
        self.account = random.choice(self.accounts)
        headers = {
            "authorization": "Bearer",
            "user-agent": self.userAgent
        }
        body = {
            "username": self.account["login"],
            "password": self.account["password"],
            "grant_type": "password"
        }
        async with aiohttp.request("POST", url, headers=headers, data=body) as response:
            responseJson = await response.json()
        return responseJson
    

    def writeToExcel(self, data):
        headers = self.analogOutputHeaders
        for row in data:
            # if any(row["type"] == ) 
            if row["type"] == "original":
                headers = self.originalOutputHeaders
                break
            
        writer = pd.ExcelWriter(self.outputFilePath, engine='openpyxl')
        df = pd.DataFrame(data)
        
        df.to_excel(writer, encoding="utf-8", index=False, header=headers)
        writer.save()
        writer.close()

    
    def readAccountsFile(self):
        rawData = pd.read_excel(self.accountsFilePath, index_col=False)
        accounts = rawData.to_numpy()
        
        accountsData = []
        for account in accounts:
            accountsData.append({
                "login": account[0],
                "password": account[1]
            })
        return accountsData


    def readDetailsFile(self):
        rawData = pd.read_excel(self.inputFilePath, header=None)
        data = [{"detailName": name, "detailNumber": number} for name, number in rawData.values.tolist()]
        del data[0]
        return data


    def run(self):
        self.accounts = self.readAccountsFile()
        self.data = self.readDetailsFile()
        if self.accounts:
            loop = asyncio.get_event_loop()
            loop.run_until_complete(self.startParsing())
        else:
            print("Введите логин и пароль от аккаунта в файл accounts.xlsx")

    def setup(self):
        isFirstLaunch = False
        # input dir
        if not os.path.exists(self.inputDirectoryPath):
            os.mkdir(self.inputDirectoryPath)
            isFirstLaunch = True
        
        # input file
        if not os.path.exists(self.inputFilePath):
            wb = Workbook()
            for i, header in enumerate(self.inputHeaders):
                wb.worksheets[0].cell(1 , i+1, value=header)
            wb.save(self.inputFilePath)
            isFirstLaunch = True


        # output dir
        if not os.path.exists(self.outputDirectoryPath):
            os.mkdir(self.outputDirectoryPath)
            isFirstLaunch = True

        # accounts file
        if not os.path.exists(self.accountsFilePath):
            wb = Workbook()
            for i, header in enumerate(self.accountsHeaders):
                wb.worksheets[0].cell(1 , i+1, value=header)
            wb.save(self.accountsFilePath)
            isFirstLaunch = True

        return isFirstLaunch

        


if __name__ == "__main__":    
    detailParser = DetailParser()
    if not detailParser.setup():
        detailParser.run()
    else:
        print("Рабочие файлы и папки созданы, заполните их!")
